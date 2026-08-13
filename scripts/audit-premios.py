# -*- coding: utf-8 -*-
"""
Auditoría del catálogo de premios contra el Excel de la campaña.

    npm run audit:premios

Vuelve a leer `recursos/premios/Calendario de Premios 2026.xlsx`, cruza sus
cuatro hojas entre sí y las compara contra el catálogo implementado en
`src/mocks/prizes.ts`. No modifica nada: sólo informa y devuelve código 1 si
algo no cierra.

Requiere Python 3 con openpyxl (`pip install openpyxl`). Es una herramienta de
desarrollo: no participa del build ni del sitio publicado.
"""
import os
import re
import sys
from collections import Counter

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instalalo con: pip install openpyxl")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "recursos", "premios", "Calendario de Premios 2026.xlsx")
CATALOG = os.path.join(ROOT, "src", "mocks", "prizes.ts")
ASSETS = os.path.join(ROOT, "src", "assets", "prizes")

EXPECTED_TYPES = 19
EXPECTED_UNITS = 89

# El Excel mezcla mayúsculas griegas con latinas (ΜΟΝΟΡΑΤΙN, ΒΟΥ, ΜΚΡ): restos
# de una conversión de codificación. Se normalizan para poder comparar.
GREEK2LAT = str.maketrans("ΑΒΕΖΗΙΚΜΝΟΡΤΥΧ", "ABEZHIKMNOPTYX")

problems = []
report = []


def check(ok, line):
    report.append(("OK " if ok else "XX ") + line)
    if not ok:
        problems.append(line)


def norm(value):
    return re.sub(r"[^A-Z0-9]", "", str(value).translate(GREEK2LAT).upper())


# --------------------------------------------------------------- Excel ------
wb = openpyxl.load_workbook(XLSX, data_only=True)

cal = [r for r in wb["CALENDARIO"].iter_rows(min_row=5, values_only=True) if r[0] is not None]
cal_count = Counter(norm(r[3]) for r in cal)
fechas = sorted(r[1] for r in cal)
check(len(cal) == EXPECTED_UNITS, f"CALENDARIO: {len(cal)} eventos (esperados {EXPECTED_UNITS})")
check(len(cal_count) == EXPECTED_TYPES, f"CALENDARIO: {len(cal_count)} tipos (esperados {EXPECTED_TYPES})")
report.append(f"    ventana: {fechas[0].date()} → {fechas[-1].date()}")

web = [(str(r[1]).strip(), r[2]) for r in wb["NOMBRE WEB"].iter_rows(min_row=4, values_only=True) if r[1]]
check(len(web) == EXPECTED_TYPES, f"NOMBRE WEB: {len(web)} tipos (esperados {EXPECTED_TYPES})")
check(sum(q for _, q in web) == EXPECTED_UNITS, f"NOMBRE WEB: {sum(q for _, q in web)} unidades (esperadas {EXPECTED_UNITS})")

prem = [str(r[2]).strip() for r in wb["PREMIOS"].iter_rows(min_row=4, values_only=True) if r[2]]
check(len(prem) == EXPECTED_UNITS, f"PREMIOS: {len(prem)} filas (esperadas {EXPECTED_UNITS})")
check(len(Counter(norm(p) for p in prem)) == EXPECTED_TYPES, f"PREMIOS: {len(Counter(norm(p) for p in prem))} tipos (esperados {EXPECTED_TYPES})")

cant = [(str(r[1]).strip(), r[2]) for r in wb["CANTIDADES"].iter_rows(min_row=4, values_only=True) if r[1]]
check(len(cant) == EXPECTED_TYPES, f"CANTIDADES: {len(cant)} tipos (esperados {EXPECTED_TYPES})")
check(sum(q for _, q in cant) == EXPECTED_UNITS, f"CANTIDADES: {sum(q for _, q in cant)} unidades (esperadas {EXPECTED_UNITS})")

# NOMBRE WEB contra CALENDARIO, premio por premio.
descuadres = [
    (name, qty, cal_count.get(norm(name)))
    for name, qty in web
    if cal_count.get(norm(name)) != qty
]
check(not descuadres, f"NOMBRE WEB vs CALENDARIO: {len(web) - len(descuadres)}/{len(web)} cantidades coinciden")
for name, qty, got in descuadres:
    report.append(f"    {name}: NOMBRE WEB {qty} vs CALENDARIO {got}")

# CANTIDADES sólo audita totales: los nombres difieren a propósito entre hojas.
web_qty = sorted(q for _, q in web)
cant_qty = sorted(q for _, q in cant)
check(web_qty == cant_qty, "CANTIDADES vs NOMBRE WEB: mismo reparto de cantidades")

# ------------------------------------------------------------- Catálogo -----
src = open(CATALOG, encoding="utf-8").read()
imports = dict(re.findall(r"import\s+(\w+)\s+from\s+'\.\./assets/prizes/([^']+)'", src))
entries = re.findall(
    r"\{\s*id:\s*'([^']+)',\s*name:\s*'([^']+)',\s*article:\s*'([^']+)',\s*"
    r"quantity:\s*(\d+),\s*image:\s*(\w+),\s*thumb:\s*(\w+),",
    src,
)
check(len(entries) == EXPECTED_TYPES, f"Catálogo: {len(entries)} premios (esperados {EXPECTED_TYPES})")

ids = [e[0] for e in entries]
check(len(set(ids)) == len(ids), f"Catálogo: {len(set(ids))} IDs únicos")
check(len({e[1] for e in entries}) == len(entries), "Catálogo: nombres únicos")
check(sum(int(e[3]) for e in entries) == EXPECTED_UNITS, f"Catálogo: {sum(int(e[3]) for e in entries)} unidades (esperadas {EXPECTED_UNITS})")

# Cantidad de cada premio contra NOMBRE WEB, en el orden de la hoja.
web_quantities = [q for _, q in web]
cat_quantities = [int(e[3]) for e in entries]
check(
    web_quantities == cat_quantities,
    "Catálogo vs NOMBRE WEB: cantidad por premio, una por una",
)
if web_quantities != cat_quantities:
    for (name, qty), entry in zip(web, entries):
        if qty != int(entry[3]):
            report.append(f"    {entry[0]}: catálogo {entry[3]} vs NOMBRE WEB {qty} ({name})")

# Cada import apunta a un archivo que existe.
faltantes = [f for f in imports.values() if not os.path.exists(os.path.join(ASSETS, f))]
check(not faltantes, f"Imágenes: {len(imports)} imports, {len(imports) - len(faltantes)} existen en disco")
for f in faltantes:
    report.append(f"    falta: src/assets/prizes/{f}")

# Nombres visibles sin códigos internos ni restos de codificación.
sucios = [e[1] for e in entries if re.search(r"MKP\d|[ΑΒΕΖΗΙΚΜΝΟΡΤΥΧ]|\d{4,}-\d", e[1])]
check(not sucios, "Nombres visibles: sin códigos internos ni caracteres griegos")
for n in sucios:
    report.append(f"    revisar: {n}")

# Artículo definido en los 19.
check(all(e[2] for e in entries), "Artículo gramatical: definido en los 19")

print("\n".join(report))
print()
if problems:
    print(f"AUDITORÍA FALLIDA: {len(problems)} problema(s).")
    sys.exit(1)
print(f"AUDITORÍA OK — {EXPECTED_TYPES} tipos, {EXPECTED_UNITS} unidades, imágenes completas.")
